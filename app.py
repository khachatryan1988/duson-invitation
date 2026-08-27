import os
import re
import hmac

from io import BytesIO
from datetime import datetime, timezone

from flask import (
    Flask,
    request,
    render_template,
    redirect,
    url_for,
    session,
    abort,
    send_file,
    flash,
)

from flask_sqlalchemy import SQLAlchemy

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter


# =========================================================
# TRANSLATIONS
# =========================================================

from translations import TRANSLATIONS

from translations_memorandum import (
    TRANSLATIONS as MEMORANDUM_TRANSLATIONS
)


# =========================================================
# EVENTS
# =========================================================

EVENT_MAIN = "baghramyan-main"
EVENT_MEMORANDUM = "memorandum-signing"


EVENT_NAMES = {
    EVENT_MAIN: "Baghramyan Residence",
    EVENT_MEMORANDUM: "Memorandum Signing",
}


# =========================================================
# APP
# =========================================================

app = Flask(__name__)

app.config["SECRET_KEY"] = os.getenv(
    "SECRET_KEY",
    "dev-change-this-secret"
)

app.config["MAX_CONTENT_LENGTH"] = 1 * 1024 * 1024


# =========================================================
# DATABASE
# =========================================================

database_url = os.getenv(
    "DATABASE_URL",
    "sqlite:///duson_event.db"
)

if database_url.startswith("postgres://"):
    database_url = database_url.replace(
        "postgres://",
        "postgresql://",
        1
    )

app.config["SQLALCHEMY_DATABASE_URI"] = database_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)


# =========================================================
# MODEL
# =========================================================

class Guest(db.Model):

    __tablename__ = "guests"

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    # =====================================================
    # SOURCE EVENT
    # =====================================================
    #
    # baghramyan-main
    # memorandum-signing
    #
    # По этому полю определяем,
    # с какой страницы пришла регистрация.
    #

    event_slug = db.Column(
        db.String(100),
        nullable=True,
        index=True
    )

    first_name = db.Column(
        db.String(100),
        nullable=False
    )

    last_name = db.Column(
        db.String(100),
        nullable=False
    )

    # Оставляем NOT NULL для совместимости
    # с существующей PostgreSQL.
    company = db.Column(
        db.String(200),
        nullable=False
    )

    position = db.Column(
        db.String(200),
        nullable=True
    )

    phone = db.Column(
        db.String(50),
        nullable=False,
        index=True
    )

    email = db.Column(
        db.String(200),
        nullable=False,
        index=True
    )


    # =====================================================
    # LEGACY FIELDS
    # =====================================================

    attendance_type = db.Column(
        db.String(30),
        nullable=False,
        default="solo"
    )

    companion_first_name = db.Column(
        db.String(100),
        nullable=True
    )

    companion_last_name = db.Column(
        db.String(100),
        nullable=True
    )

    companion_company = db.Column(
        db.String(200),
        nullable=True
    )

    companion_position = db.Column(
        db.String(200),
        nullable=True
    )

    companion_phone = db.Column(
        db.String(50),
        nullable=True
    )

    companion_email = db.Column(
        db.String(200),
        nullable=True
    )


    # =====================================================
    # OTHER
    # =====================================================

    special_notes = db.Column(
        db.Text,
        nullable=True
    )

    consent = db.Column(
        db.Boolean,
        nullable=False,
        default=False
    )

    created_at = db.Column(
        db.DateTime,
        nullable=False,
        default=lambda: datetime.now(timezone.utc)
    )


# =========================================================
# LANGUAGES
# =========================================================

SUPPORTED_LANGUAGES = (
    "hy",
    "ru",
    "en",
)


def normalize_language(lang):

    if lang not in SUPPORTED_LANGUAGES:
        return "hy"

    return lang


# =========================================================
# VALIDATION MESSAGES
# =========================================================

ERROR_MESSAGES = {

    "hy": {

        "first_name":
            "Խնդրում ենք լրացնել անունը։",

        "last_name":
            "Խնդրում ենք լրացնել ազգանունը։",

        "company":
            "Խնդրում ենք լրացնել ընկերության / "
            "կազմակերպության անվանումը։",

        "position":
            "Խնդրում ենք լրացնել պաշտոնը։",

        "phone":
            "Խնդրում ենք լրացնել ճիշտ հեռախոսահամար։",

        "email":
            "Խնդրում ենք լրացնել ճիշտ էլեկտրոնային հասցե։",

        "consent":
            "Անհրաժեշտ է համաձայնել տվյալների "
            "օգտագործման պայմաններին։",

        "already":
            "Այս հեռախոսահամարով կամ էլեկտրոնային "
            "հասցեով գրանցում արդեն առկա է։",
    },


    "ru": {

        "first_name":
            "Пожалуйста, укажите имя.",

        "last_name":
            "Пожалуйста, укажите фамилию.",

        "company":
            "Пожалуйста, укажите компанию / организацию.",

        "position":
            "Пожалуйста, укажите должность.",

        "phone":
            "Пожалуйста, укажите корректный номер телефона.",

        "email":
            "Пожалуйста, укажите корректный адрес электронной почты.",

        "consent":
            "Необходимо согласиться с условиями "
            "использования предоставленных данных.",

        "already":
            "Регистрация с таким номером телефона "
            "или электронной почтой уже существует.",
    },


    "en": {

        "first_name":
            "Please enter your first name.",

        "last_name":
            "Please enter your last name.",

        "company":
            "Please enter your company / organization.",

        "position":
            "Please enter your position.",

        "phone":
            "Please enter a valid phone number.",

        "email":
            "Please enter a valid email address.",

        "consent":
            "You must agree to the data usage terms.",

        "already":
            "A registration with this phone number "
            "or email address already exists.",
    },
}


# =========================================================
# HELPERS
# =========================================================

def get_translation(lang):

    lang = normalize_language(lang)

    return TRANSLATIONS[lang]


def get_memorandum_translation(lang):

    lang = normalize_language(lang)

    return MEMORANDUM_TRANSLATIONS[lang]


def admin_required():

    if not session.get("admin_logged_in"):
        abort(403)


def normalize_phone(phone: str) -> str:

    phone = (
            phone
            or ""
    ).strip()

    phone = re.sub(
        r"[^\d+]",
        "",
        phone
    )

    if phone.startswith("00"):
        phone = "+" + phone[2:]

    return phone


def valid_phone(phone: str) -> bool:

    if not phone:
        return False

    digits = re.sub(
        r"\D",
        "",
        phone
    )

    return len(digits) >= 8


def valid_email(email: str) -> bool:

    email = (
            email
            or ""
    ).strip()

    pattern = (
        r"^[^@\s]+@[^@\s]+\.[^@\s]+$"
    )

    return bool(
        re.match(
            pattern,
            email
        )
    )


def get_event_name(event_slug: str) -> str:

    return EVENT_NAMES.get(
        event_slug,
        event_slug or ""
    )


# =========================================================
# ROOT
# =========================================================

@app.route("/")
def home():

    return redirect(
        url_for(
            "index",
            lang="hy"
        )
    )


# =========================================================
# =========================================================
# MAIN BAGHRAMYAN VERSION
# =========================================================
# =========================================================
#
# /hy
# /ru
# /en
#
# event_slug:
# baghramyan-main
#
# =========================================================


# =========================================================
# MAIN INDEX
# =========================================================

@app.route("/<lang>")
def index(lang):

    lang = normalize_language(
        lang
    )

    t = get_translation(
        lang
    )

    return render_template(
        "index.html",
        t=t,
        lang=lang
    )


# =========================================================
# MAIN REGISTRATION
# =========================================================
#
# REQUIRED:
#
# first_name
# last_name
# company
# position
# phone
# email
# consent
#
# =========================================================

@app.route(
    "/<lang>/register",
    methods=["POST"]
)
def register(lang):

    lang = normalize_language(
        lang
    )

    t = get_translation(
        lang
    )

    messages = ERROR_MESSAGES[
        lang
    ]


    # =====================================================
    # FORM DATA
    # =====================================================

    first_name = request.form.get(
        "first_name",
        ""
    ).strip()

    last_name = request.form.get(
        "last_name",
        ""
    ).strip()

    company = request.form.get(
        "company",
        ""
    ).strip()

    position = request.form.get(
        "position",
        ""
    ).strip()

    phone = normalize_phone(
        request.form.get(
            "phone",
            ""
        )
    )

    email = request.form.get(
        "email",
        ""
    ).strip().lower()

    special_notes = request.form.get(
        "special_notes",
        ""
    ).strip()

    consent = (
            request.form.get(
                "consent"
            )
            == "on"
    )


    # =====================================================
    # VALIDATION
    # =====================================================

    errors = []


    if not first_name:
        errors.append(
            messages["first_name"]
        )


    if not last_name:
        errors.append(
            messages["last_name"]
        )


    if not company:
        errors.append(
            messages["company"]
        )


    if not position:
        errors.append(
            messages["position"]
        )


    if not valid_phone(
            phone
    ):
        errors.append(
            messages["phone"]
        )


    if not valid_email(
            email
    ):
        errors.append(
            messages["email"]
        )


    if not consent:
        errors.append(
            messages["consent"]
        )


    # =====================================================
    # VALIDATION ERROR
    # =====================================================

    if errors:

        return render_template(
            "index.html",

            t=t,
            lang=lang,

            errors=errors,

            form_data=request.form

        ), 400


    # =====================================================
    # DUPLICATE CHECK
    # =====================================================
    #
    # Проверяем только внутри MAIN event.
    #
    # Один человек может зарегистрироваться:
    #
    # 1 раз на baghramyan-main
    # +
    # 1 раз на memorandum-signing
    #
    # =====================================================

    existing_guest = (
        Guest.query
        .filter(
            Guest.event_slug == EVENT_MAIN,
            db.or_(
                Guest.email == email,
                Guest.phone == phone
            )
        )
        .first()
    )


    if existing_guest:

        return redirect(
            url_for(
                "already_registered",
                lang=lang
            )
        )


    # =====================================================
    # SAVE
    # =====================================================

    guest = Guest(

        # ВАЖНО
        event_slug=EVENT_MAIN,

        first_name=first_name,

        last_name=last_name,

        company=company,

        position=position,

        phone=phone,

        email=email,

        attendance_type="solo",

        companion_first_name=None,
        companion_last_name=None,
        companion_company=None,
        companion_position=None,
        companion_phone=None,
        companion_email=None,

        special_notes=(
                special_notes
                or None
        ),

        consent=consent,
    )


    db.session.add(
        guest
    )

    db.session.commit()


    return redirect(
        url_for(
            "thank_you",
            lang=lang
        )
    )


# =========================================================
# MAIN THANK YOU
# =========================================================

@app.route(
    "/<lang>/thank-you"
)
def thank_you(lang):

    lang = normalize_language(
        lang
    )

    t = get_translation(
        lang
    )

    return render_template(
        "thank_you.html",
        t=t,
        lang=lang
    )


# =========================================================
# MAIN ALREADY REGISTERED
# =========================================================

@app.route(
    "/<lang>/already-registered"
)
def already_registered(lang):

    lang = normalize_language(
        lang
    )

    t = get_translation(
        lang
    )

    return render_template(
        "already_registered.html",
        t=t,
        lang=lang
    )


# =========================================================
# =========================================================
# MEMORANDUM VERSION
# =========================================================
# =========================================================
#
# /memorandum-signing/hy
# /memorandum-signing/ru
# /memorandum-signing/en
#
# event_slug:
# memorandum-signing
#
# =========================================================


# =========================================================
# MEMORANDUM ROOT
# =========================================================

@app.route(
    "/memorandum-signing"
)
def memorandum_home():

    return redirect(
        url_for(
            "memorandum_index",
            lang="hy"
        )
    )


# =========================================================
# MEMORANDUM INDEX
# =========================================================

@app.route(
    "/memorandum-signing/<lang>"
)
def memorandum_index(lang):

    lang = normalize_language(
        lang
    )

    t = get_memorandum_translation(
        lang
    )

    return render_template(
        "memorandum/index.html",

        t=t,
        lang=lang
    )


# =========================================================
# MEMORANDUM REGISTRATION
# =========================================================
#
# REQUIRED:
#
# first_name
# last_name
# consent
#
# OPTIONAL:
#
# company
# position
# phone
# email
# special_notes
#
# =========================================================

@app.route(
    "/memorandum-signing/<lang>/register",
    methods=["POST"]
)
def memorandum_register(lang):

    lang = normalize_language(
        lang
    )

    t = get_memorandum_translation(
        lang
    )

    messages = ERROR_MESSAGES[
        lang
    ]


    # =====================================================
    # FORM DATA
    # =====================================================

    first_name = request.form.get(
        "first_name",
        ""
    ).strip()

    last_name = request.form.get(
        "last_name",
        ""
    ).strip()

    company = request.form.get(
        "company",
        ""
    ).strip()

    position = request.form.get(
        "position",
        ""
    ).strip()

    phone = normalize_phone(
        request.form.get(
            "phone",
            ""
        )
    )

    email = request.form.get(
        "email",
        ""
    ).strip().lower()

    special_notes = request.form.get(
        "special_notes",
        ""
    ).strip()

    consent = (
            request.form.get(
                "consent"
            )
            == "on"
    )


    # =====================================================
    # VALIDATION
    # =====================================================

    errors = []


    # REQUIRED
    if not first_name:

        errors.append(
            messages["first_name"]
        )


    # REQUIRED
    if not last_name:

        errors.append(
            messages["last_name"]
        )


    # OPTIONAL PHONE
    #
    # Пустой разрешён.
    # Если заполнен — проверяем формат.
    #

    if phone and not valid_phone(
            phone
    ):

        errors.append(
            messages["phone"]
        )


    # OPTIONAL EMAIL
    #
    # Пустой разрешён.
    # Если заполнен — проверяем формат.
    #

    if email and not valid_email(
            email
    ):

        errors.append(
            messages["email"]
        )


    # REQUIRED CONSENT
    if not consent:

        errors.append(
            messages["consent"]
        )


    # =====================================================
    # VALIDATION ERROR
    # =====================================================

    if errors:

        return render_template(
            "memorandum/index.html",

            t=t,
            lang=lang,

            errors=errors,

            form_data=request.form

        ), 400


    # =====================================================
    # DUPLICATE CHECK
    # =====================================================
    #
    # Пустые phone/email не проверяем.
    #
    # Ищем дубликат ТОЛЬКО среди
    # memorandum-signing.
    #
    # =====================================================

    duplicate_conditions = []


    if email:

        duplicate_conditions.append(
            Guest.email == email
        )


    if phone:

        duplicate_conditions.append(
            Guest.phone == phone
        )


    existing_guest = None


    if duplicate_conditions:

        existing_guest = (
            Guest.query
            .filter(
                Guest.event_slug == EVENT_MEMORANDUM,
                db.or_(
                    *duplicate_conditions
                )
            )
            .first()
        )


    if existing_guest:

        return redirect(
            url_for(
                "memorandum_already_registered",
                lang=lang
            )
        )


    # =====================================================
    # SAVE
    # =====================================================

    guest = Guest(

        # ВАЖНО
        event_slug=EVENT_MEMORANDUM,

        first_name=first_name,

        last_name=last_name,

        # В существующей БД company NOT NULL.
        company=(
                company
                or ""
        ),

        position=(
                position
                or None
        ),

        # phone/email в существующей БД NOT NULL.
        phone=(
                phone
                or ""
        ),

        email=(
                email
                or ""
        ),

        attendance_type="solo",

        companion_first_name=None,

        companion_last_name=None,

        companion_company=None,

        companion_position=None,

        companion_phone=None,

        companion_email=None,

        special_notes=(
                special_notes
                or None
        ),

        consent=consent,
    )


    db.session.add(
        guest
    )

    db.session.commit()


    # =====================================================
    # SUCCESS
    # =====================================================

    return redirect(
        url_for(
            "memorandum_thank_you",
            lang=lang
        )
    )


# =========================================================
# MEMORANDUM THANK YOU
# =========================================================

@app.route(
    "/memorandum-signing/<lang>/thank-you"
)
def memorandum_thank_you(lang):

    lang = normalize_language(
        lang
    )

    t = get_memorandum_translation(
        lang
    )

    return render_template(
        "memorandum/thank_you.html",

        t=t,
        lang=lang
    )


# =========================================================
# MEMORANDUM ALREADY REGISTERED
# =========================================================

@app.route(
    "/memorandum-signing/<lang>/already-registered"
)
def memorandum_already_registered(lang):

    lang = normalize_language(
        lang
    )

    t = get_memorandum_translation(
        lang
    )

    return render_template(
        "memorandum/already_registered.html",

        t=t,
        lang=lang
    )


# =========================================================
# ADMIN LOGIN
# =========================================================

@app.route(
    "/admin/login",
    methods=[
        "GET",
        "POST"
    ]
)
def admin_login():

    if request.method == "GET":

        return render_template(
            "admin_login.html"
        )


    password = request.form.get(
        "password",
        ""
    )


    admin_password = os.getenv(
        "ADMIN_PASSWORD",
        "change-me"
    )


    if not hmac.compare_digest(
            password,
            admin_password
    ):

        flash(
            "Неверный пароль",
            "error"
        )

        return render_template(
            "admin_login.html"
        ), 403


    session[
        "admin_logged_in"
    ] = True


    return redirect(
        url_for(
            "admin_guests"
        )
    )


# =========================================================
# ADMIN LOGOUT
# =========================================================

@app.route(
    "/admin/logout"
)
def admin_logout():

    session.clear()

    return redirect(
        url_for(
            "admin_login"
        )
    )


# =========================================================
# ADMIN PANEL
# =========================================================

@app.route(
    "/admin"
)
def admin_guests():

    if not session.get(
            "admin_logged_in"
    ):

        return redirect(
            url_for(
                "admin_login"
            )
        )


    guests = (
        Guest.query
        .order_by(
            Guest.created_at.desc()
        )
        .all()
    )


    registration_count = len(
        guests
    )


    main_count = sum(
        1
        for guest in guests
        if guest.event_slug == EVENT_MAIN
    )


    memorandum_count = sum(
        1
        for guest in guests
        if guest.event_slug == EVENT_MEMORANDUM
    )


    companions = 0

    total_people = (
        registration_count
    )


    return render_template(
        "admin.html",

        guests=guests,

        registration_count=registration_count,

        main_count=main_count,

        memorandum_count=memorandum_count,

        companions=companions,

        total_people=total_people,

        event_names=EVENT_NAMES,
    )


# =========================================================
# EXCEL EXPORT
# =========================================================

@app.route(
    "/admin/export"
)
def export_excel():

    admin_required()


    guests = (
        Guest.query
        .order_by(
            Guest.created_at.asc()
        )
        .all()
    )


    # =====================================================
    # WORKBOOK
    # =====================================================

    wb = Workbook()

    ws = wb.active

    ws.title = "Baghramyan Registrations"


    # =====================================================
    # HEADERS
    # =====================================================

    headers = [

        "ID",

        "Անուն",

        "Ազգանուն",

        "Գրանցման աղբյուր",

        "Ընկերություն / կազմակերպություն",

        "Պաշտոն",

        "Հեռախոսահամար",

        "Էլեկտրոնային հասցե",

        "Հատուկ նշումներ",

        "Համաձայնություն",

        "Գրանցման ամսաթիվ",
    ]


    ws.append(
        headers
    )


    # =====================================================
    # HEADER STYLE
    # =====================================================

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="252A24"
    )


    header_font = Font(
        color="FFFFFF",
        bold=True
    )


    thin = Side(
        style="thin",
        color="D8D8D2"
    )


    for cell in ws[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        cell.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )


    ws.row_dimensions[1].height = 38


    # =====================================================
    # DATA
    # =====================================================

    for guest in guests:

        created = guest.created_at


        if (
                created
                and created.tzinfo
        ):

            created = (
                created
                .astimezone(
                    timezone.utc
                )
                .replace(
                    tzinfo=None
                )
            )


        ws.append([

            guest.id,

            guest.first_name,

            guest.last_name,

            # Откуда пришла регистрация
            get_event_name(
                guest.event_slug
            ),

            guest.company or "",

            guest.position or "",

            guest.phone or "",

            guest.email or "",

            guest.special_notes or "",

            (
                "Այո"
                if guest.consent
                else "Ոչ"
            ),

            (
                created.strftime(
                    "%d.%m.%Y %H:%M"
                )
                if created
                else ""
            ),
            ])


    # =====================================================
    # WIDTHS
    # =====================================================

    widths = [

        8,      # ID
        20,     # First name
        20,     # Last name
        26,     # Event
        32,     # Company
        25,     # Position
        20,     # Phone
        35,     # Email
        45,     # Notes
        18,     # Consent
        23,     # Date
    ]


    for index, width in enumerate(
            widths,
            start=1
    ):

        ws.column_dimensions[
            get_column_letter(
                index
            )
        ].width = width


    # =====================================================
    # BODY STYLE
    # =====================================================

    for row in ws.iter_rows(
            min_row=2
    ):

        for cell in row:

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )


    # =====================================================
    # EXCEL OPTIONS
    # =====================================================

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = (
        ws.dimensions
    )


    # =====================================================
    # SAVE
    # =====================================================

    output = BytesIO()

    wb.save(
        output
    )

    output.seek(
        0
    )


    filename = (
        "Baghramyan_registration_"
        f"{datetime.now().strftime('%Y-%m-%d')}"
        ".xlsx"
    )


    return send_file(

        output,

        as_attachment=True,

        download_name=filename,

        mimetype=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )


# =========================================================
# HEALTH CHECK
# =========================================================

@app.route(
    "/health"
)
def health():

    return {
        "status": "ok"
    }


# =========================================================
# LOCAL DEVELOPMENT
# =========================================================

if __name__ == "__main__":

    app.run(

        host="0.0.0.0",

        port=int(
            os.getenv(
                "PORT",
                "5000"
            )
        ),

        debug=(
                os.getenv(
                    "FLASK_DEBUG",
                    "0"
                )
                == "1"
        )
    )